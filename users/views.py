from django.shortcuts import render
from rest_framework.permissions import IsAuthenticated
from datetime import datetime
from django.utils.translation import gettext_lazy as _

# Create your views here.
from knox.views import LoginView as KnoxLoginView
from knox.views import LogoutView
from django.shortcuts import render
from rest_framework.fields import empty
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
from .models import *
from rest_framework import generics
from .serializers import *
from rest_framework.response import Response
from django.contrib.auth import SESSION_KEY, authenticate, login
from django.core.mail import send_mail
from datetime import date, datetime
from datetime import datetime

from rest_framework import permissions

import string
import random
import pdb
from django.core.mail import send_mail
from django.template.loader import render_to_string
import openpyxl

# s = date.today()
class CreateUser(generics.CreateAPIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def post(self, request, format=None):
        result = {}
        result['status'] = 'NOK'
        result['valid'] = False
        result["result"] = {'message': 'Unauthorized', 'data': []}
        if request.user.is_authenticated:

            serializer = UserSerializer(data=request.data)
            if serializer.is_valid():

                try:
                    username = serializer.validated_data['email']
                    # password = ''.join(
                    #     random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k=8))
                    password='123456'
                    serializer.validated_data['password'] = password
                    serializer.save()
                    name = request.data['name']
                    email = request.data['email']
                    msg_plain = ''
                    msg_html = render_to_string(
                        'users/create_user.html', {'name': name, 'username': username, 'password': password})
                    # send_mail('CRM Dashboard', msg_plain,
                    #             'info@sorasorimukhyomontri.com', [email], html_message=msg_html,)
                    send_mail('CRM Dashboard', msg_plain,
                              'info@pkconnect.com', [email], html_message=msg_html, )

                    # if request.data['user_role']=='4':
                    #     obj=User.objects.filter(user_role=4).order_by('-analyst_id')[0]
                    #     obj1=User.objects.last()
                    #     User.objects.filter(id=obj1.id).update(analyst_id=obj.analyst_id+1)

                except:
                    result['status'] = 'NOK'
                    result['valid'] = False
                    result['result']['message'] = "Error in sending mail"
                    return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

                # result =dict()
                result['status'] = 'OK'
                result['valid'] = True
                result['result']['message'] = "User created successfully !"
                return Response(result, status=status.HTTP_200_OK)
            else:
                result['result']['message'] = (list(serializer.errors.keys())[
                                                   0] + ' - ' + list(serializer.errors.values())[0][0]).capitalize()
                return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        return Response(result, status=status.HTTP_401_UNAUTHORIZED)

class Login(KnoxLoginView):
    permission_classes = (permissions.AllowAny,)
    serializer_class = LoginSerializer

    def post(self, request, format=None):
        serializer = LoginSerializer(data=request.data)
        result = {}
        result['status'] = 'NOK'
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if serializer.is_valid():
            try:
                user_data = authenticate(email=serializer.validated_data['email'],
                                         password=serializer.validated_data['password'])

            except:
                # Response data
                result['status'] = 'NOK'
                result['valid'] = False
                result['result']['message'] = 'User not present'
                # Response data
                return Response(result, status=status.HTTP_204_NO_CONTENT)

            if user_data is not None:
                user_details = User.objects.all().filter(email=user_data).values('id', 'name', 'email', 'phone',
                                                                                 'registered_on', 'emp_code',
                                                                                 'user_role_id', 'current_status',
                                                                                 'is_active')
                # print(user_details)
                if user_data.is_active:
                    login(request, user_data)
                    data = super(Login, self).post(request)
                    data = data.data
                    print(data)
                    # data['message'] = "Login successfully"
                    data['user_info'] = user_details

                # Response data
                result['status'] = "OK"
                result['valid'] = True
                result['result']['message'] = "Login successfully"
                result['result']['data'] = data
                # result['result']['data'] = data
                # Response data
                return Response(result, status=status.HTTP_200_OK)
            else:

                # Response data
                result['status'] = "NOK"
                result['valid'] = False
                result['result']['message'] = 'Invalid Credentials'
                # Response data
                return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
        # Response data
        result['status'] = "NOK"
        result['valid'] = False
        result['result']['message'] = (
                    list(serializer.errors.keys())[0] + ' - ' + list(serializer.errors.values())[0][0]).capitalize()
        # Response data
        return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

class Logoutview(LogoutView):
    # permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        result = {}
        result['status'] = 'NOK'
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}
        if request.user.is_authenticated:
            try:
                request._auth.delete()
            except:
                # Response data
                result['status'] = "NOK"
                result['valid'] = False
                result['result']['message'] = 'Error while logging out'
                # Response data
                return Response(result, status=status.HTTP_200_OK)
            # Response data
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = 'Logout successfully !'
            # Response data
            return Response(result, status=status.HTTP_200_OK)

class ChangePassword(APIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = ChangePasswordSerializer

    def post(self, request, format=None):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:

            if request.user.is_anonymous:
                result['result']['message'] = "User Invalid"
                return Response(result, status=status.HTTP_200_OK)

            serializer = ChangePasswordSerializer(data=request.data)
            if serializer.is_valid():
                user = request.user

                if user.check_password(serializer.data['old_password']):
                    new_password = serializer.data['new_password']
                    user.set_password(serializer.data['new_password'])
                    request.user.save()

                    # Response data
                    result['status'] = "OK"
                    result['valid'] = True
                    result['result']['message'] = "Password Changed"
                    # Response data
                    return Response(result, status=status.HTTP_201_CREATED)
                else:
                    result['result']['message'] = "Pasword did not match"
                    return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

class ResetPasswordLink(APIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = ResetPasswordLinkSerializer

    def post(self, request, format=None):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}
        key = ''.join(random.choices(
            string.ascii_uppercase + string.digits, k=20))
        serializer_data = dict()
        serializer_data['key'] = key
        serializer_data['time'] = datetime.now()

        user_data = User.objects.all().filter(email=request.data['email'])

        if len(user_data) == 0:
            result['result']['message'] = 'Enter a valid email address.'
            return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
        else:
            user_id = user_data.values()[0]['id']

            try:
                prev_data = ResetPassword.objects.filter(user_id_id=user_id)
                len_prev = len(prev_data.values())

            except:
                pass

            if len_prev > 0:

                serializer_data = {}
                serializer_data['key'] = key
                serializer_data['time'] = datetime.now()
                # serializer_data['user_id']  = user_id
                serializer = ResetPasswordLinkSerializer(ResetPassword.objects.get(
                    user_id_id=user_id), data=serializer_data, partial=True)

                if serializer.is_valid():
                    serializer.save()
                else:
                    result['result']['message'] = (
                                list(serializer.errors.keys())[0] + ' - ' + list(serializer.errors.values())[0][
                            0]).capitalize()
                    return Response(result, status=status.HTTP_200_OK)

            else:

                serializer_data = {}
                serializer_data['key'] = key
                serializer_data['time'] = datetime.now()
                serializer_data['user_id'] = user_id

                serializer = ResetPasswordLinkSerializer(data=serializer_data)

                if serializer.is_valid():
                    serializer.save()
                else:
                    result['result']['message'] = (
                                list(serializer.errors.keys())[0] + ' - ' + list(serializer.errors.values())[0][
                            0]).capitalize()
                    return Response(result, status=status.HTTP_200_OK)

        try:
            # data = "https://sorasorimukhyomontri.com/ssm-dashboard/auth/password/create/" + key
            data = "https://youthclub.pkconnect.com/stage/crm_dashboard/auth/password/create/" + key
            email = request.data['email']
            obj = User.objects.filter(email=email).values()
            obj = obj[0]['name']
            msg_plain = ''
            msg_html = render_to_string(
                'users/resetpass.html', {'link': data, 'obj': obj})

            # send_mail('Reset Password request', msg_plain, 'info@sorasorimukhyomontri.com',
            #           [request.data['email']], html_message=msg_html, )
            send_mail('Reset Password request', msg_plain, 'info@pkconnect.com',
                      [request.data['email']], html_message=msg_html, )
        except:
            result['result']['message'] = "Error while sending mail"
            return Response(result, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # Response data
        result['status'] = "OK"
        result['valid'] = True
        result['result']['message'] = "Reset password link has been sent to your mail address"
        # Response data

        return Response(result, status=status.HTTP_200_OK)

class UserResetPassword(APIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = ResetPasswordsSerializer

    def post(self, request, format=None):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}
        if request.data['password'] != request.data['confirm_password']:
            result['result']['message'] = "Both passwords must be same"
            return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        try:
            user_details = ResetPassword.objects.filter(key=request.data['key']).values('id', 'key', 'time',
                                                                                        'user_id_id__email',
                                                                                        'user_id_id')
            if len(user_details) == 0:
                result['result']['message'] = "Invalid user"
                return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
        except:
            pass

        email_id = user_details[0]['user_id_id__email']
        start_time = user_details[0]['time']
        endtime = datetime.now()
        duration = (endtime - start_time).total_seconds() / 60.0

        if duration > 20:
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Request Time Out!"
            return Response(result, status=status.HTTP_408_REQUEST_TIMEOUT)

        serializer = ResetPasswordsSerializer(data=request.data)

        try:
            user = User.objects.get(email=email_id)
        except:
            result['result']['message'] = "Invalid Email Address"
            return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        if serializer.is_valid():
            user.set_password(serializer.data['password'])
            user.save()
            obj = User.objects.filter(email=email_id).values()
            obj = obj[0]['name']
            msg_plain = ''
            msg_html = render_to_string('users/reset.html', {'name': obj})
            # send_mail('Password Reset  Sucessful', msg_plain,
            #           'info@sorasorimukhyomontri.com', [email_id], html_message=msg_html, )
            send_mail('Password Reset  Sucessful', msg_plain,
                      'info@pkconnect.com', [email_id], html_message=msg_html, )

            prev_data = ResetPassword.objects.get(id=user_details[0]['id'])

            serializer_data = {}
            serializer_data['key'] = ""
            serializer = ResetPasswordLinkSerializer(
                prev_data, data=serializer_data, partial=True)
            if serializer.is_valid():
                serializer.save()
            else:
                result['status'] = "NOK"
                result['valid'] = False
                result['result']['message'] = (
                            list(serializer.errors.keys())[0] + ' - ' + list(serializer.errors.values())[0][
                        0]).capitalize()
                return Response(result, status=status.HTTP_200_OK)

            # Response data
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Password reset successfully !"
            # Response data
            return Response(result, status=status.HTTP_200_OK)
        else:
            result['result']['message'] = (
                        list(serializer.errors.keys())[0] + ' - ' + list(serializer.errors.values())[0][0]).capitalize()
            return Response(result, status=status.HTTP_400_BAD_REQUEST)

class GetUsersData(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def get(self, request):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:
            users_data = User.objects.all().exclude(is_deleted=1).values()
            # Response data
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Data fetched successfully"
            result['result']['data'] = users_data
            # Response data
            return Response(result, status=status.HTTP_200_OK)

    def post(self, request, format=None):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:
            users = request.data['user_id']

            User.objects.filter(id__in=users).delete()
            # Response data
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "User deleted successfully !"
            # Response data
            return Response(result, status=status.HTTP_200_OK)

class GetChips(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:
            try:
                request.data['search_string']
                print(request.data['search_string'])
            except:
                result['result']['message'] = " Search parameter is missing"
                return Response(result, status=status.HTTP_204_NO_CONTENT)

            search_string = request.data['search_string']
            start_date = request.data['start_date']
            end_date = request.data['end_date']
            active_status = request.data['status']
            qs = User.objects.filter(is_deleted=1).values_list('id')
            data = User.objects.all().exclude(id=(request.user.id))
            data = data.exclude(id__in=qs)

            if start_date and end_date:
                data = data.filter(registered_on__range=[
                    start_date, end_date]).all()
            elif start_date != "":
                data = data.filter(registered_on__gte=start_date).all()
            elif end_date != "":
                data = data.filter(registered_on__lte=end_date).all()

            if active_status != "":
                data = data.filter(is_active=active_status).all()

            All = data.filter(name__icontains=search_string).count()
            Super_Admin = data.filter(
                name__icontains=search_string, user_role_id=1).count()
            LAU_Admin = data.filter(
                name__icontains=search_string, user_role_id=2).count()
            LAU_Consent_Agent = data.filter(
                name__icontains=search_string, user_role_id__in=[3,4]).count()
            LAU_WA_Agent = data.filter(
                name__icontains=search_string, user_role_id__in=[3,4]).count()
            CAU_Admin = data.filter(
                name__icontains=search_string, user_role_id=5).count()
            CAU_Agent = data.filter(
                name__icontains=search_string, user_role_id=6).count()
            COU_Admin = data.filter(
                name__icontains=search_string, user_role_id=7).count()
            COU_Agent = data.filter(
                name__icontains=search_string, user_role_id=8).count()

            result1 = {}
            result1['buckets'] = [{"bucket_id": "0", "bucket_name": "All", "value": All},
                                  {"bucket_id": "1", "bucket_name": "Super Admin", "value": Super_Admin},
                                  {"bucket_id": "2", "bucket_name": "LAU Admin", "value": LAU_Admin},
                                  {"bucket_id": "3", "bucket_name": "LAU Consent Agent", "value": LAU_Consent_Agent},
                                  {"bucket_id": "4", "bucket_name": "LAU WA Agent", "value": LAU_WA_Agent},
                                  {"bucket_id": "5", "bucket_name": "CAU Admin", "value": CAU_Admin},
                                  {"bucket_id": "6", "bucket_name": "CAU Agent", "value": CAU_Agent},
                                  {"bucket_id": "7", "bucket_name": "COU Admin", "value": COU_Admin},
                                  {"bucket_id": "8", "bucket_name": "COU Agent", "value": COU_Agent},
                                  ]
            # Response data
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Data fetched successfully"
            result['result']['data'] = result1
            # Response data
            return Response(result, status=status.HTTP_200_OK)

class GetChipsData(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:
            bucket_id = request.data['bucket_id']
            search_string = request.data['search_string']
            # error handle if bucket is empty
            start_date = request.data['start_date']
            end_date = request.data['end_date']
            active_status = request.data['status']

            if bucket_id == "":
                result['result']['message'] = "Chip ID cannot be empty"
                return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
            if int(bucket_id) > 8:
                result['result']['message'] = "Role Id not available"
                return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
            qs = User.objects.filter(is_deleted=1).values_list('id')
            # data = User.objects.all().exclude(id=(request.user.id))
            data = User.objects.all().exclude(id__in=qs)
            # data = data.exclude(id__in=qs)
            if start_date and end_date:
                data = data.filter(registered_on__range=[
                    start_date, end_date]).all()
            elif start_date != "":
                data = data.filter(registered_on__gte=start_date).all()
            elif end_date != "":
                data = data.filter(registered_on__lte=end_date).all()

            if active_status != "":
                data = data.filter(is_active=active_status).all()

            '''bucket data'''
            if (bucket_id == '0'):
                data = data.all().filter(name__icontains=search_string).values('id', 'name', 'email', 'phone',
                                                                               'registered_on', 'emp_code',
                                                                               'user_role_id__role_name',
                                                                               'current_status', 'is_active',)

            elif (bucket_id == '3' or bucket_id == '4'):
                data = data.all().filter(name__icontains=search_string, user_role_id__in=[3,4]).values('id', 'name',
                                                                                                       'email', 'phone',
                                                                                                       'registered_on',
                                                                                                       'emp_code',
                                                                                                       'user_role_id__role_name',
                                                                                                       'current_status',
                                                                                                       'is_active')
            # if (bucket_id == '0' or bucket_id == '1' or bucket_id == '2' or bucket_id == '5' or bucket_id == '6' or bucket_id == '7' or bucket_id == '8'):
            else:
                data = data.all().filter(name__icontains=search_string, user_role_id=bucket_id).values('id', 'name',
                                                                                                       'email', 'phone',
                                                                                                       'registered_on',
                                                                                                       'emp_code',
                                                                                                       'user_role_id__role_name',
                                                                                                       'current_status',
                                                                                                       'is_active')

            paginated_data = self.paginate_queryset(data, request, view=self)
            paginated_data = self.get_paginated_response(paginated_data).data

            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Data fetched successfully"
            result['result']['next'] = paginated_data['next']
            result['result']['previous'] = paginated_data['previous']
            result['result']['count'] = paginated_data['count']
            result['result']['data'] = paginated_data['results']

            return Response(result, status=status.HTTP_200_OK)
        return Response(result, status=status.HTTP_401_UNAUTHORIZED)

class UpdateUserAndPermissions(APIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def patch(self, request, format=None):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}
        # pdb.set_trace()
        if request.user.is_authenticated:
            user_id = request.data['id']

            if (not request.user.is_anonymous):
                user = User.objects.get(id=user_id)

                serializer = EditUserSerializer(
                    user, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    # Response data
                    result['status'] = "OK"
                    result['valid'] = True
                    result['result']['message'] = "User updated successfully !"

                    return Response(result, status=status.HTTP_200_OK)
                else:
                    result['result']['message'] = (
                                list(serializer.errors.keys())[0] + ' - ' + list(serializer.errors.values())[0][
                            0]).capitalize()
                    return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

            else:
                result['result']['message'] = str(request.user) + " : Unauthorized user"
                return Response(result, status=status.HTTP_401_UNAUTHORIZED)
        return Response(result, status=status.HTTP_401_UNAUTHORIZED)

class DeleteUser(APIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def post(self, request, format=None):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:
            if request.data['ids'] == "":
                result['result']['message'] = "ids cannot be empty"
                return Response(result, status=status.HTTP_401_UNAUTHORIZED)
            else:
                ids = request.data['ids'].split(",")
                qs = User.objects.filter(id__in=ids)
                for obj in qs:
                    temp = User.objects.filter(email=obj).values()
                    temp.update(is_active=False)
                    temp.update(is_deleted=True)

            # Response data
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Records deleted successfully"
            # Response data
            return Response(result, status=status.HTTP_200_OK)

class UpdateUserStatus(APIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def post(self, request, format=None):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:
            if request.data['ids'] == "":
                result['status'] = "OK"
                result['valid'] = False
                result['result']['message'] = "ids cannot be empty"
                return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

            if request.data['status'] == "":
                result['result']['message'] = "status cannot be empty"
                return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

            user_ids = request.data['ids']
            user_ids = user_ids.split(",")

            if len(user_ids) > 0:
                data = User.objects.filter(id__in=user_ids)
                for i in data:
                    i.is_active = request.data['status']

                User.objects.bulk_update(data, ['is_active'])
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Status updated successfully !"

            return Response(result, status=status.HTTP_200_OK)

class GetRoles(APIView):
    def post(self, request):
        result = {}
        result['status'] = "NOK"
        result['valid'] = False
        result['result'] = {"message": "Unauthorized access", "data": []}

        if request.user.is_authenticated:
            data = Roles.objects.all().values()
            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Data fetched successfully !"
            result['result']['data'] = data

            return Response(result, status=status.HTTP_200_OK)

        return Response(result, status=status.HTTP_200_OK)

class CreateBulkUsers(generics.CreateAPIView):
    serializer_class = UserSerializer

    def post(self, request, format=None):
        result = {}
        result['status'] = 'NOK'
        result['valid'] = False
        result["result"] = {'message': 'Unauthorized'}
        if request.user.is_authenticated:

            workbook = openpyxl.load_workbook(request.data["file"])
            sheet = workbook['Sheet1']

            # added_list = []
            notadded_list = []
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                record = {}
                record["emp_code"] = row[1]
                record["name"] = row[2]
                record["phone"] = row[3]
                record["email"] = row[4]
                record["user_role"] = 8
                record["acs"] = ""
                record["priority_level"] = ""
                record["rm_code"] = ""
                record["assigned_analyst"] = ""
                serializer = UserSerializer(data=record)

                if serializer.is_valid():
                    serializer.validated_data['password'] = "123456"
                    # added_list.append(record)
                    serializer.save()
                else:
                    notadded_list.append(record)
            workbook.close()

            result['status'] = "OK"
            result['valid'] = True
            result['result']['message'] = "Status updated successfullly"
            # result['result']['added'] = added_list
            result['result']['not_added'] = notadded_list
            return Response(result, status=status.HTTP_200_OK)
        return Response(result, status=status.HTTP_401_UNAUTHORIZED)
